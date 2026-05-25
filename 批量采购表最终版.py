import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook

MUNICIPALITIES = {'北京市', '上海市', '天津市', '重庆市'}

# ======================================================
# 地址去重清洗（🌟 终极优化：处理镇/街道交替叠加的情况）
# ======================================================
def clean_duplicate_address(addr):
    if not isinstance(addr, str) or str(addr).strip() in ['nan', '']:
        return ""
    
    addr = str(addr).strip()
    
    # 1. 消除带有不同行政区划后缀的嵌套叠加（如：广东省中山市南朗街道 广东省中山市南朗镇）
    prov, city = "", ""
    if addr[:3] in MUNICIPALITIES:
        prov = addr[:3]
    else:
        match = re.match(r'^(.{2,8}?(?:省|自治区))(.{2,8}?(?:市|自治州|地区|盟))?', addr)
        if match:
            prov = match.group(1)
            city = match.group(2) or ""
            
    prefix = prov + city
    if prefix:
        # 寻找前25个字符内是否出现了第二次“省+市”组合
        second_idx = addr.find(prefix, len(prov))
        if second_idx > 0 and second_idx <= 25: 
            middle = addr[len(prefix):second_idx]
            
            # 情况 A：[省市][街道][省市][镇] -> 保留前面的街道，剔除后面的省市镇
            if re.match(r'^.{1,10}?(?:区|县|镇|乡|街道|旗|村)$', middle):
                match_second = re.match(r'^' + re.escape(prefix) + r'.{1,10}?(?:区|县|镇|乡|街道|旗|村)', addr[second_idx:])
                if match_second:
                    addr = addr[:second_idx] + addr[second_idx + len(match_second.group(0)):]
                else:
                    addr = addr[:second_idx] + addr[second_idx + len(prefix):]
            
            # 情况 B：[省市][省市][镇] -> 剔除开头的省市
            elif middle == "":
                addr = addr[second_idx:]
                
    # 2. 消除完全一致的超级叠词
    n = len(addr)
    changed = True
    while changed:
        changed = False
        for l in range(n // 2, 2, -1):
            sub = addr[:l]
            if addr[l:].startswith(sub):
                addr = addr[l:]
                n = len(addr)
                changed = True
                break
                
    # 3. 消除中间部分的局部连续叠加
    changed = True
    while changed:
        changed = False
        n = len(addr)
        for i in range(n):
            for l in range(3, (n - i) // 2 + 1): 
                sub = addr[i:i+l]
                if addr[i+l:i+2*l] == sub:
                    addr = addr[:i+l] + addr[i+2*l:]
                    changed = True
                    break
            if changed:
                break
                
    return addr

# ======================================================
# 地址解析（🌟 新增：直筒子市独立逻辑处理）
# ======================================================
def parse_address(addr):

    addr = str(addr).strip()

    if addr[:3] in MUNICIPALITIES:
        province = addr[:3]
        pe = 3

    elif '自治区' in addr:
        pe = addr.index('自治区') + 3
        province = addr[:pe]

    elif '省' in addr:
        pe = addr.index('省') + 1
        province = addr[:pe]

    else:
        province = addr[:3]
        pe = 3

    try:
        ce = addr.index('市', pe) + 1
    except:
        ce = pe

    city = addr[pe:ce]

    if not city:
        city = province

    endings = ['区', '县', '镇', '乡', '旗', '街道']

    de = len(addr)

    for ch in endings:
        try:
            pos = addr.index(ch, ce) + len(ch)
            if pos < de:
                de = pos
        except:
            pass

    try:
        pos = addr.index('市', ce) + 1
        if pos < de:
            de = pos
    except:
        pass

    district = addr[ce:de]
    detail = addr[de:]

    for prefix in [addr[:de], city + district, district]:
        if prefix and detail.startswith(prefix):
            detail = detail[len(prefix):]
            break

    # 🌟 特判：处理中山市、东莞市等无区县的“直筒子市”
    if city in ['中山市', '东莞市', '嘉峪关市', '儋州市'] and district != city:
        if district:
            # 把原本被切作区县的镇/街道，放回详细地址里
            detail = district + detail
        # 强制将区县列设为市名
        district = city

    return province, city, district, detail


# ======================================================
# 获取当前目录（兼容 PyInstaller）
# ======================================================

if getattr(sys, 'frozen', False):
    CURRENT_DIR = os.path.dirname(sys.executable)
else:
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))


# ======================================================
# 自动寻找文件
# ======================================================

path_12 = None
if os.path.exists(os.path.join(CURRENT_DIR, "12.xlsx")):
    path_12 = os.path.join(CURRENT_DIR, "12.xlsx")

path_123 = None
if os.path.exists(os.path.join(CURRENT_DIR, "123.xlsx")):
    path_123 = os.path.join(CURRENT_DIR, "123.xlsx")

if not path_12 or not path_123:
    print("\n【错误】缺少xlsx文件！")
    print("请确保：\n12.xlsx\n123.xlsx\n在同一个文件夹内")
    input("\n按回车退出...")
    sys.exit()

output_path = os.path.join(CURRENT_DIR, "123_output.xlsx")


# ======================================================
# 读取12
# ======================================================

print("=================== 第一步：读取12数据 ===================")

df_12 = pd.read_excel(path_12, dtype=str)
df_12.columns = df_12.columns.str.strip()
df_12 = df_12.fillna('')

if '收货地址' in df_12.columns:
    df_12['收货地址'] = df_12['收货地址'].str.replace('广西壮族自治区', '广西省')
    print("已将源数据中的'广西壮族自治区'全局替换为'广西省'")
    
    # 运行强力去重逻辑
    df_12['收货地址'] = df_12['收货地址'].apply(clean_duplicate_address)
    print("已全局智能清理地址中的连续重复冗余信息")

print("12表字段：")
print(list(df_12.columns))


# ======================================================
# 打开123模板并重构列位置
# ======================================================

print("\n=================== 第二步：重构123模板列位置 ===================")

wb = load_workbook(path_123)
ws = wb.active

current_headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        current_headers[str(val).strip()] = col

anchor_idx = current_headers.get('收货地址（复制）')
if not anchor_idx:
    anchor_idx = current_headers.get('收货地址')

if anchor_idx:
    print(f"检测到锚点列在第 {anchor_idx} 列，准备进行结构调整...")
    ws.cell(1, anchor_idx).value = '收货地址（复制）'

    if '收货省份' not in current_headers:
        print(f"正在原列（第 {anchor_idx} 列）的正左侧自动插入 4 个解析列...")
        ws.insert_cols(anchor_idx, 4)
        ws.cell(1, anchor_idx).value = '收货省份'
        ws.cell(1, anchor_idx + 1).value = '收货城市'
        ws.cell(1, anchor_idx + 2).value = '收货区县'
        ws.cell(1, anchor_idx + 3).value = '收货地址'
        print("【成功】4个解析列已精确插入至左侧！")
    else:
        print("【提示】解析列已存在，跳过插入操作。")
else:
    print("【警告】未检测到收货地址列！")


headers_123 = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header:
        headers_123[str(header).strip()] = col

print("\n更新后的123表字段：")
print(list(headers_123.keys()))


# ======================================================
# 行数校验与清理
# ======================================================

excel_data_rows = ws.max_row - 1
target_data_count = len(df_12)

print(f"\n[行数统计] 12表：{target_data_count} 行 | 123模板：{excel_data_rows} 行")

if excel_data_rows > target_data_count:
    start_delete_row = target_data_count + 2
    rows_to_delete = excel_data_rows - target_data_count
    print(f"⚠️ 正在清理多余旧数据...")
    ws.delete_rows(idx=start_delete_row, amount=rows_to_delete)


# ======================================================
# 自动复制字段
# ======================================================

print("\n=================== 第三步：复制字段 ===================")

copy_map = {
    '订单号': '其它出库业务单号',
    '收货人': '收货人',
    '收货电话': '收货电话',
    '收货地址': '收货地址（复制）',
    'SKU采购总⾦额（含税）': '单价',
    '采购数量（采购单位）': '数量',
    'SKU编码': 'SKU编码',
}

for source_col, target_col in copy_map.items():
    if source_col not in df_12.columns:
        print(f"【警告】12缺少字段：{source_col}")
        continue
    if target_col not in headers_123:
        print(f"【警告】123缺少字段：{target_col}")
        continue

    target_excel_col = headers_123[target_col]

    for i in range(target_data_count):
        excel_row = i + 2
        value = df_12.iloc[i][source_col]

        if target_col in ['SKU编码', '其它出库业务单号', '收货电话']:
            value = str(value).strip()
            if value.endswith('.0'):
                value = value[:-2]

        ws.cell(excel_row, target_excel_col).value = value

    print(f"【成功】{source_col} -> {target_col}")


# ======================================================
# 地址解析与备注
# ======================================================

print("\n=================== 第四步：地址解析与备注生成 ===================")

if '收货地址' in df_12.columns:

    parsed_data = []
    for addr in df_12['收货地址']:
        if str(addr).strip():
            parsed_data.append(parse_address(addr))
        else:
            parsed_data.append(("", "", "", ""))

    parsed_df = pd.DataFrame(
        parsed_data,
        columns=['收货省份', '收货城市', '收货区县', '收货地址']
    )

    clean_addresses = (
        df_12['收货地址']
        .astype(str)
        .fillna('')
        .str.replace('nan', '')
        .str.strip()
    )

    remark_series = (
        df_12['收货人'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + df_12['收货电话'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + clean_addresses
    ).str.strip()

    parsed_df['备注'] = remark_series

    for field in parsed_df.columns:
        if field not in headers_123:
            new_col = ws.max_column + 1
            ws.cell(1, new_col).value = field
            headers_123[field] = new_col
            print(f"【自动新增列】{field}")

        target_excel_col = headers_123[field]

        for i in range(target_data_count):
            excel_row = i + 2
            value = parsed_df.iloc[i][field]
            ws.cell(excel_row, target_excel_col).value = value

        print(f"【成功】写入：{field}")


# ======================================================
# 保存
# ======================================================

print("\n=================== 第五步：保存文件 ===================")

wb.save(output_path)

print("\n【✨ 全部完成 ✨】")
print("4个解析列已自动插入")
print("已完成广西壮族自治区 -> 广西省 替换")
print(f"已写入 {target_data_count} 行数据")
print(f"输出文件：{output_path}")

input("\n按回车退出...")
